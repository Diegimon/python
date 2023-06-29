import string
import random
from openpyxl import Workbook, load_workbook
import pandas as pd
import os
import subprocess

# Verificar e instalar bibliotecas ausentes
required_libraries = ['openpyxl']

for library in required_libraries:
    try:
        __import__(library)
    except ImportError:
        print(f"A biblioteca {library} não está instalada. Instalando...")
        subprocess.check_call(['pip', 'install', library])


reset = '\033[0m'
red = '\033[0;49;91m'
green = '\033[32m'
yellow = '\033[33m'
blue = '\033[34m'
ciano = '\033[36m'
planilha = None


def criar_planilha():
    # Verifica se a planilha já existe
    if os.path.exists("planilha.xlsx"):
        print("Planilha already exists")
        return

    # Cria uma nova planilha

    # Cria um DataFrame com alguns dados de exemplo
    dados = {'name': ['Pedro Fernández', 'Juan Pérez', 'Pepito Limón', 'María Gamesa'],
             'password': [123, 321, 123, 321],
             'account': ['PD201819HASRTR07', 'JP201819HASRTR06', 'PL211819HASRTR06', 'MG211811MASRTR05'],
             'balance': [1, 1, 1, 1]}
    df_main = pd.DataFrame(dados)
    df_main = df_main.reset_index()

    # Define a coluna 'name' como índice para a planilha 'main'
    df_main.set_index('name', inplace=True)

    # Cria um DataFrame vazio para a planilha 'history'
    df_history = pd.DataFrame(columns=['account', 'register'])

    # Preenche a coluna 'account' da planilha 'history' com os mesmos dados da coluna 'account' na planilha 'main'
    df_history['account'] = df_main['account']

    # Cria a coluna 'register' na planilha 'history' com valores vazios
    df_history['register'] = 'Account created'

    # Cria um arquivo Excel e adiciona as planilhas
    with pd.ExcelWriter("planilha.xlsx") as writer:
        df_main.to_excel(writer, sheet_name='main', index=True)
        df_history.to_excel(writer, sheet_name='history', index=False)

    print(f"{green}Planilhas created successfully.{reset}")


def create_account():
    df_main = pd.read_excel("planilha.xlsx", sheet_name='main')
    df_history = pd.read_excel("planilha.xlsx", sheet_name='history')

    while True:
        name = input("Account Name: ")
        if name in df_main['name'].values:
            print(f"{red}Account already exists{reset}")
        else:
            while True:
                password = input("Password: ")
                confirm = input("Confirm Password: ")
                if confirm != password:
                    print(f"{red}The passwords are incorrect{reset}")
                else:
                    balance = 1
                    account = gerar_sequencia()
                    nova_conta_main = pd.DataFrame({'name': [name],
                                                    'password': [password],
                                                    'account': [account],
                                                   'balance': [balance]})
                    nova_conta_history = pd.DataFrame({'account': [account],
                                                       'register': ['Account created']})

                    # Adiciona a nova conta ao DataFrame 'main'
                    df_main = pd.concat(
                        [df_main, nova_conta_main], ignore_index=True)

                    # Adiciona a nova linha ao DataFrame 'history'
                    df_history = pd.concat(
                        [df_history, nova_conta_history], ignore_index=True)

                    # Salva os DataFrames atualizados no arquivo Excel
                    with pd.ExcelWriter("planilha.xlsx") as writer:
                        df_main.to_excel(
                            writer, sheet_name='main', index=False)
                        df_history.to_excel(
                            writer, sheet_name='history', index=False)

                    print(f"{green}Account created successfully{reset}")
                    break

            break


def gerar_sequencia():
    # Gera uma sequência aleatória de duas letras
    letras = random.choices(string.ascii_uppercase, k=2)

    # Gera um número de 6 dígitos
    numero_6_digitos = random.randint(100000, 999999)

    # Gera um número aleatório de 1 a 9
    numero_1_a_9 = random.randint(1, 9)

    # Retorna a sequência gerada
    return ''.join(letras) + str(numero_6_digitos) + "RASTR0" + str(numero_1_a_9)


def validate_login(name, password):
    planilha = load_workbook("planilha.xlsx")
    planilha_ativa = planilha['main']
    user = False
    valida = False
    account_name = None
    account_number = None
    balance = None

    for linha in planilha_ativa.iter_rows(min_row=2, values_only=True):
        if name == linha[0]:
            user = True
            print(f"{green}Valid user{reset}")
            if password == linha[2]:
                print(f"{green}Valid password{reset}")
                valida = True
                account_name = linha[0]
                account_number = linha[3]
                balance = linha[4]
                break

    if not user:
        print(f"{red}Invalid username{reset}")

    elif not valida:
        print(f"{red}Invalid password{reset}")

    return valida, account_name, account_number, balance


def deposit_amount(account_number, deposit):
    planilha = load_workbook("planilha.xlsx")
    main = planilha["main"]
    number = f"{account_number}"
    encontrado = False

    for celula in main["D"]:
        if number == celula.value:
            encontrado = True
            linha = celula.row
            balance = main[f'E{linha}'].value
            balance += deposit
            main[f"E{linha}"] = balance
            print(f"New Balance: {balance}")
            print(f"{green}Deposit successfully{reset}")
            reg = f"Deposit {deposit}$ New balance: {balance}"
            planilha.save("planilha.xlsx")

            break

    if not encontrado:
        print("Não encontrado")
    planilha = load_workbook("planilha.xlsx")
    aba_history = planilha['history']
    coluna_a = aba_history['A']

    for celula in coluna_a:
        if celula.value == account_number:
            linha = celula.row
            new_history = aba_history[f'B{linha}'].value + "/" + reg
            aba_history[f'B{linha}'].value = new_history
            planilha.save("planilha.xlsx")


def show_history(account_number):

    planilha = load_workbook("planilha.xlsx")
    history = planilha['history']
    for number in history['A']:
        if account_number == number.value:
            linha = number.row
            print(yellow, history[f"B{linha}"].value, reset)


def withdrawal_amount(account_number, withdraw):
    planilha = load_workbook("planilha.xlsx")
    main = planilha["main"]
    number = f"{account_number}"
    encontrado = False

    for celula in main["D"]:
        if number == celula.value:
            encontrado = True
            linha = celula.row
            balance = main[f'E{linha}'].value
            if withdraw > balance:
                print(f"{red}You don't have enough!{reset}")
            balance -= withdraw
            main[f"E{linha}"] = balance
            print(f"New Balance: {balance}")
            print(f"{green}withdrawal successfully{reset}")

            # Atualizando historico

            reg = f" {withdraw}$ withdrawn, New balance: {balance}"
            break

    if not encontrado:
        print(f"{red}Não encontrado{reset}")
    planilha.save("planilha.xlsx")
    aba_history = planilha['history']
    coluna_a = aba_history['A']

    for celula in coluna_a:
        if celula.value == account_number:
            linha = celula.row
            new_history = aba_history[f'B{linha}'].value + "/" + reg
            aba_history[f'B{linha}'].value = new_history
            planilha.save("planilha.xlsx")


def delete_account(account_number):
    planilha = load_workbook("planilha.xlsx")
    main = planilha['main']
    find = False
    for linha in main['D']:
        if account_number == linha.value:
            linha_da_conta = linha.row
            main.delete_rows(linha_da_conta)
            planilha.save("planilha.xlsx")
            print(f"{yellow}Account deleted!{reset}")
            find = True
    if find == False:
        print("Account not found!")
    reg = "Account deleted"
    aba_history = planilha['history']
    coluna_a = aba_history['A']

    for celula in coluna_a:
        if celula.value == account_number:
            linha = celula.row
            new_history = aba_history[f'B{linha}'].value + "/" + reg
            aba_history[f'B{linha}'].value = new_history
            planilha.save("planilha.xlsx")


def menu():
    print("=" * 20)
    valid = ["1", "2", "3", "4", "5", "6", "", " "]
    m = input(
        f"{blue}[Enter] To Close{reset}\n[1] Create Account\n[2] Login\n[3] Deposit\n[4] Withdrawal\n[5] Help\n\n: ")
    while m not in valid:
        print("Invalid option!")
        m = input("[1] Create Account\n[2] Login\n[3] Close\n: ")

    return m

# Função principal


def main():
    criar_planilha()
    while True:
        m = menu()
        if m == '1':
            create_account()
        elif m == '2':
            print("Login to your account")
            name = str(input("Name: "))
            password = str(input("Password: "))

            valida, account_name, account_number, balance = validate_login(
                name, password)

            if valida == True:
                account_number = f"{account_number}"
                print(f"{yellow}Login successful!")
                print("Account Information:")
                print("Name:", account_name)
                print("Account:", account_number)
                print("Balance:", balance, reset)

                while True:
                    print("=" * 20)
                    menu_option = input(
                        f"{ciano}[1] Deposit\n[2] Withdrawal\n[3] History\n[4] Delete\n[5] Close\n: {reset}")

                    if menu_option == '1':
                        deposit = float(input("Deposit amount: "))
                        deposit_amount(account_number, deposit)
                    elif menu_option == '2':
                        withdraw = float(input("Withdraw amount: "))
                        withdrawal_amount(account_number, withdraw)
                    elif menu_option == '3':
                        show_history(account_number)
                    elif menu_option == '4':
                        delete_account(account_number)
                        break
                    elif menu_option == '5':
                        break
                    else:
                        print("Invalid option!")
            else:
                print(f"{red}Login failed!.{reset}")

        elif m == '3':
            account_number = input("Account number: ")
            planilha = load_workbook("planilha.xlsx")
            planilha_ativa = planilha['main']
            user = False
            for linha in planilha_ativa["D"]:
                if account_number == linha.value:
                    user = True
            if user == True:
                cash = float(input("Value: "))
                deposit_amount(account_number, cash)

            else:
                print("Account not found")

        elif m == '4':
            account_number = input("Account number: ")
            planilha = load_workbook("planilha.xlsx")
            planilha_ativa = planilha['main']
            user = False
            for linha in planilha_ativa["D"]:
                if account_number == linha.value:
                    user = True
                    value = planilha_ativa.cell(row=linha.row, column=5).value

            if user == True:
                cash = float(input("Value: "))
                if cash <= value:
                    withdrawal_amount(account_number, cash)
                else:
                    print("Insufficient funds")

        elif m == '5':
            print(f"""{blue}You can get help by contacting the link below:
https://api.whatsapp.com/send?hone=996528343

or via email: dscontac@hotmail.com

code: github.com/diegimon{reset}""")
        # elif m == '6':
        elif m == '' or m == ' ':
            print(
                f"Prog{green}ram finis{blue}hed th{yellow}anks {reset}for the pre{yellow}fer{red}ence!{reset}")
            finaly = input("[press any button to end]\n")
            break


main()
