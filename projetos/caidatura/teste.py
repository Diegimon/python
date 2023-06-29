import string
import random
from openpyxl import Workbook, load_workbook


def withdrawal_amount(account_number, withdraw):
    planilha = load_workbook("planilha.xlsx")
    main = planilha["main"]
    number = f"{account_number}"
    encontrado = False

    for celula in main["D"]:
        if number == celula.value:
            encontrado = True
            linha = celula.row
            balance = main.cell(row=linha, column=5).value
            if withdraw > balance:
                print("You don't have enough!")
            balance -= withdraw
            main.cell(row=linha, column=5).value = balance
            print(f"New Balance: {balance}")
            print("Withdrawal successfully")

            # Atualizando historico
            reg = f" {withdraw}$ withdrawn, New balance: {balance}"
            break
