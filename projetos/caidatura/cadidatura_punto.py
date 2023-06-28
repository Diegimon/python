import string
import random
from openpyxl import Workbook, load_workbook
import pandas as pd
from random import randint
import os
import subprocess

# Verificar e instalar bibliotecas ausentes
required_libraries = ['openpyxl']

for library in required_libraries:
    try:
        __import__(library)
    except ImportError:
        print(f"A biblioteca {library} não está instalada. Instalando...")
        subprocess.check_call(['pip', 'install', library])

# A partir deste ponto, você pode importar as bibliotecas normalmente

reset = '\033[0m'
red = '\033[0;49;91m'
green = '\033[32m'
yellow = '\033[33m'
blue = '\033[34m'
ciano = '\033[36m'
planilha = None

# Função para exibir o menu


def menu():
    valid = ["1", "2", "3"]
    m = input("[1] Create Account\n[2] Login\n[3] Close\n: ")
    while m not in valid:
        print(f"{red}Invalid option{reset}")
        m = input("[1] Create Account\n[2] Login\n[3] Close\n: ")
    return m


# Função para criar a planilha
def criar_planilha():
    # Verifica se a planilha já existe
    if os.path.exists("planilha.xlsx"):
        print("Planilha already exists")
        return

    # Cria uma nova planilha

    # Cria um DataFrame com alguns dados de exemplo
    dados = {'name': ['Pedro Fernández', 'Juan Pérez', 'Pepito Limón', 'María Gamesa'],
             'password': [123, 321, 123, 321],
             'account': ['PD201819HASRTR07', 'JP201819HASRTR06', 'PL211819HASRTR06', 'MG211811MASRTR05']}
    df_main = pd.DataFrame(dados)
    df_main = df_main.reset_index()

    # Define a coluna 'name' como índice para a planilha 'main'
    df_main.set_index('name', inplace=True)

    # Cria um DataFrame vazio para a planilha 'history'
    df_history = pd.DataFrame(columns=['account', 'register'])

    # Preenche a coluna 'account' da planilha 'history' com os mesmos dados da coluna 'account' na planilha 'main'
    df_history['account'] = df_main['account']

    # Cria a coluna 'register' na planilha 'history' com valores vazios
    df_history['register'] = 'Account created'

    # Cria um arquivo Excel e adiciona as planilhas
    with pd.ExcelWriter("planilha.xlsx") as writer:
        df_main.to_excel(writer, sheet_name='main', index=True)
        df_history.to_excel(writer, sheet_name='history', index=False)

    # Imprime o DataFrame 'main' com o índice
    print(df_main)

    print("Planilhas created successfully.")


# Função para criar uma conta

def create_account():
    df_main = pd.read_excel("planilha.xlsx", sheet_name='main')
    df_history = pd.read_excel("planilha.xlsx", sheet_name='history')

    while True:
        name = input("Account Name: ")
        if name in df_main['name'].values:
            print("Account already exists")
        else:
            while True:
                password = input("Password: ")
                confirm = input("Confirm Password: ")
                if confirm != password:
                    print("The passwords are incorrect")
                else:
                    account = gerar_sequencia()
                    nova_conta_main = pd.DataFrame({'name': [name],
                                                    'password': [password],
                                                    'account': [account]})
                    nova_conta_history = pd.DataFrame({'account': [account],
                                                       'register': ['Account created']})

                    # Adiciona a nova conta ao DataFrame 'main'
                    df_main = pd.concat(
                        [df_main, nova_conta_main], ignore_index=True)

                    # Adiciona a nova linha ao DataFrame 'history'
                    df_history = pd.concat(
                        [df_history, nova_conta_history], ignore_index=True)

                    # Salva os DataFrames atualizados no arquivo Excel
                    with pd.ExcelWriter("planilha.xlsx") as writer:
                        df_main.to_excel(
                            writer, sheet_name='main', index=False)
                        df_history.to_excel(
                            writer, sheet_name='history', index=False)

                    print("Account created successfully")
                    break

            break


def gerar_sequencia():
    # Gera uma sequência aleatória de duas letras
    letras = random.choices(string.ascii_uppercase, k=2)

    # Gera um número de 6 dígitos
    numero_6_digitos = random.randint(100000, 999999)

    # Gera um número aleatório de 1 a 9
    numero_1_a_9 = random.randint(1, 9)

    # Retorna a sequência gerada
    return ''.join(letras) + str(numero_6_digitos) + "RASTR0" + str(numero_1_a_9)


def validate_login(name, password):
    planilha = load_workbook("planilha.xlsx")
    planilha_ativa = planilha['main']
    valida = False
    account_name = None
    account_number = None
    balance = None

    for linha in range(2, planilha_ativa.max_row + 1):
        if planilha_ativa[f'D{linha}'].value == name:
            print("Valid user")
            if planilha_ativa[f'C{linha}'].value == password:
                print("Valid password")
                valida = True
                account_name = planilha_ativa[f'A{linha}'].value
                account_number = planilha_ativa[f'D{linha}'].value
                balance = 1
            else:
                print("Invalid password")
        else:
            print("Invalid username")

    return valida, account_name, account_number, balance


def update_history(account_number, reg):
    planilha = load_workbook("planilha.xlsx")
    aba_history = planilha['history']
    coluna_a = aba_history['A']

    # Itera sobre as células da coluna A
    for celula in coluna_a:
        if celula.value == account_number:
            linha = celula.row
            aba_history[f'B{linha}'].value += "-" + reg
            print(aba_history[f'B{linha}'].value)

    # Salva as alterações no arquivo
    planilha.save("planilha.xlsx")


def deposit_amount(account_number):
    planilha = load_workbook("planilha.xlsx")
    planilha_ativa = planilha['Sheet']

    linhas_atualizadas = []
    for linha in planilha_ativa.iter_rows(min_row=2, values_only=True):
        current_account_number = linha[2]  # Número da conta na planilha

        if current_account_number == account_number:
            current_balance = linha[3]  # Saldo atual da conta
            deposit = float(input("Enter the deposit amount: "))
            new_balance = current_balance + deposit

            linha_atualizada = list(linha)
            linha_atualizada[3] = new_balance  # Atualiza o saldo na linha
            linhas_atualizadas.append(linha_atualizada)
        else:
            linhas_atualizadas.append(list(linha))

    # Remove todas as linhas existentes
    planilha_ativa.delete_rows(2, planilha_ativa.max_row)

    # Adiciona as linhas atualizadas
    for linha_atualizada in linhas_atualizadas:
        planilha_ativa.append(linha_atualizada)
    reg = str((f'deposit of {deposit}$ made'))
    update_history(account_number, reg)

    planilha.save("planilha.xlsx")

    print(f"{green}Deposit successfully!{reset}")


# Função para realizar um saque
def withdrawal_amount(account_number):
    planilha = load_workbook("planilha.xlsx")
    planilha_ativa = planilha['Sheet']

    linhas_atualizadas = []
    for linha in planilha_ativa.iter_rows(min_row=2, values_only=True):
        current_account_number = linha[2]  # Número da conta na planilha

        if current_account_number == account_number:
            current_balance = linha[3]  # Saldo atual da conta
            withdrawal = float(input("Enter the withdrawal amount: "))
            if withdrawal > current_balance:
                print("Insufficient funds!")
                return
            new_balance = current_balance - withdrawal

            linha_atualizada = list(linha)
            linha_atualizada[3] = new_balance  # Atualiza o saldo na linha
            linhas_atualizadas.append(linha_atualizada)
        else:
            linhas_atualizadas.append(list(linha))

    # Remove todas as linhas existentes
    planilha_ativa.delete_rows(2, planilha_ativa.max_row)

    # Adiciona as linhas atualizadas
    for linha_atualizada in linhas_atualizadas:
        planilha_ativa.append(linha_atualizada)

    planilha.save("planilha.xlsx")
    reg = (f'Loot of {withdrawal}$ successfully carried out!')
    update_history(account_number, reg)

    print(f"{green}Withdrawal successfully!{reset}")


def delete_account(account_number):
    planilha = load_workbook("planilha.xlsx")
    planilha_ativa = planilha.active

    linhas_atualizadas = []
    for linha in planilha_ativa.iter_rows(min_row=2, values_only=True):
        current_account_number = linha[2]  # Número da conta na planilha

        if current_account_number != account_number:
            linhas_atualizadas.append(list(linha))

    # Remove todas as linhas existentes
    planilha_ativa.delete_rows(2, planilha_ativa.max_row)
    planilha.save("planilha.xlsx")
    print(f"{green}Account deleted successfully!{reset}")

# Função para exibir o menu e retornar a opção escolhida


def menu():
    valid = ["1", "2", "3"]
    m = input("[1] Create Account\n[2] Login\n[3] Close\n: ")
    while m not in valid:
        print("Invalid option!")
        m = input("[1] Create Account\n[2] Login\n[3] Close\n: ")

    return m

# Função principal


def main():
    criar_planilha()
    while True:
        m = menu()
        if m == '1':
            create_account()
        elif m == '2':
            print("Login to your account")
            name = input("Name: ")
            password = input("Password: ")

            valida, name, account_number, account_balance = validate_login(
                name, password)

            if account_number:
                print(f"{ciano}Login successful!")
                print("Account Information:")
                print("Name:", name)
                print("Account:", account_number)
                print("Balance:", account_balance, reset)

                while True:
                    print("=" * 20)
                    menu_option = input(
                        f"{yellow}[1] Deposit\n[2] Withdrawal\n[3] History\n[4] Delete\n[5] Close\n: {reset}")

                    if menu_option == '1':
                        deposit_amount(account_number)
                    elif menu_option == '2':
                        withdrawal_amount(account_number)
                    elif menu_option == '3':
                        show_history(account_number)
                    elif menu_option == '4':
                        delete_account(account_number)
                        break
                    elif menu_option == '5':
                        break
                    else:
                        print("Invalid option!")
            else:
                print(f"{red}Login failed! Invalid name or password.{reset}")
        elif m == '3':
            break


main()
